"""
NEM Generation Map builder.

Inputs (in same folder):
  - KCI Datafile Compiled NEM.xlsx
  - NEM Generation Information Oct 2025.xlsx
  - (optional, future) AEMO state map ingest

Output:
  - projects.json        unified project table
  - nem_generation_map.html   interactive Folium map

Stage logic (per brief):
  - In NEM Gen Info AND on AEMO map     -> use AEMO stage (Application/Committed)
  - In KCI AND NEM Gen Info, NOT on map -> Application
  - In KCI only                         -> Enquiry
  - NEM Gen Info Unit Status drives Existing / Committed / Anticipated buckets.
  - KCI NER type drives Enquiry vs Application when only in KCI.

We don't yet have the AEMO map extracted, so:
  - Treat NEM Gen Info "Committed" / "Anticipated" / "In Commissioning" as the
    AEMO-map "Committed" bucket.
  - Treat "Publicly Announced" as Application by default (later: if found on
    AEMO map, upgrade to Committed; if only in KCI, downgrade to Enquiry).
"""

from __future__ import annotations
import json
import re
import math
import random
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
INPUTS = ROOT / "data" / "inputs"
INTERMEDIATE = ROOT / "data" / "intermediate"
OUTPUTS = ROOT / "outputs"
INTERMEDIATE.mkdir(parents=True, exist_ok=True)
OUTPUTS.mkdir(parents=True, exist_ok=True)
KCI_FILE = INPUTS / "KCI Datafile Compiled NEM.xlsx"
NEM_FILE = INPUTS / "NEM Generation Information Oct 2025.xlsx"

REGION_TO_STATE = {
    "NSW1": "NSW", "VIC1": "VIC", "QLD1": "QLD", "SA1": "SA", "TAS1": "TAS",
}

# Approximate region centroids (lat, lon) - used until real geocoding is wired.
REGION_CENTROID = {
    "NSW1": (-32.5, 147.0),
    "VIC1": (-36.8, 144.5),
    "QLD1": (-22.5, 145.5),
    "SA1":  (-33.5, 138.6),
    "TAS1": (-42.0, 146.5),
}
REGION_JITTER = {  # half-width in degrees for the random jitter box
    "NSW1": (2.0, 3.0),
    "VIC1": (1.2, 2.2),
    "QLD1": (4.0, 4.5),
    "SA1":  (2.5, 2.5),
    "TAS1": (0.8, 0.8),
}

# AEMO Generation Information legend colours (approximate hex matches).
STAGE_COLOUR = {
    "Existing":     "#1f2937",  # dark grey / black
    "Committed":    "#16a34a",  # green
    "Anticipated":  "#f59e0b",  # amber/yellow
    "Application":  "#2563eb",  # blue
    "Enquiry":      "#dc2626",  # red
    "Withdrawn":    "#9ca3af",  # light grey
    "Unknown":      "#6b7280",
}

STAGE_ORDER = ["Existing", "Committed", "Anticipated", "Application", "Enquiry", "Withdrawn", "Unknown"]


def norm_name(s: str | None) -> str:
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"[​\xa0]", " ", s)             # zero-width / nbsp
    s = re.sub(r"\([^)]*\)", " ", s)                 # drop parenthetical
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())


def to_float(x) -> float | None:
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def read_nem_gen_info() -> list[dict]:
    wb = openpyxl.load_workbook(NEM_FILE, data_only=True, read_only=True)
    ws = wb["ExistingGeneration&NewDevs"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # header is row 2 (index 1); data from row 3
            continue
        region = row[0]
        if not region or region not in REGION_TO_STATE:
            continue
        site = row[2]
        if not site:
            continue
        unit_status = (row[14] or "").strip().rstrip("*").strip()
        # collapse the weird "Committed*" / "Committed﻿" variants
        if unit_status.lower().startswith("committed"):
            unit_status = "Committed"
        cap = to_float(row[12]) or to_float(row[11]) or to_float(row[9])
        rows.append({
            "site_name": str(site).strip(),
            "region": region,
            "state": REGION_TO_STATE[region],
            "owner": (row[3] or "").strip() if row[3] else "",
            "technology": (row[4] or "").strip() if row[4] else "",
            "fuel": (row[5] or "").strip() if row[5] else "",
            "duid": (row[6] or "").strip() if row[6] else "",
            "capacity_mw": cap,
            "storage_mwh": to_float(row[13]),
            "unit_status": unit_status,
            "asset_type": (row[1] or "").strip() if row[1] else "",
            "_source": "NEM",
            "_key": norm_name(site),
        })
    return rows


def read_kci() -> list[dict]:
    wb = openpyxl.load_workbook(KCI_FILE, data_only=True, read_only=True)
    ws = wb["Master"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # header row index 2
            continue
        region = row[12]
        if region not in REGION_TO_STATE:
            continue
        site = row[10]
        if not site:
            continue
        status = (row[6] or "").strip()
        # exclude withdrawn/cancelled per AEMO map convention
        if status.lower() in {"withdrawn", "cancelled"}:
            continue
        ner = (row[3] or "").lower()
        if "application to connect" in ner:
            kci_stage = "Application"
        elif "connection enquiry" in ner:
            kci_stage = "Enquiry"
        else:
            kci_stage = "Enquiry"
        cap_upper = to_float(row[14]) or to_float(row[25]) or to_float(row[13])
        rows.append({
            "site_name": str(site).strip().replace("​", "").replace("\xa0", " ").strip(),
            "region": region,
            "state": REGION_TO_STATE[region],
            "owner": (row[7] or "").strip() if row[7] else "",
            "technology": (row[17] or "").strip() if row[17] else "",
            "fuel": "",
            "duid": "",
            "capacity_mw": cap_upper,
            "storage_mwh": None,
            "location_desc": (row[11] or "").strip() if row[11] else "",
            "ner_type": row[3] or "",
            "activity_status": status,
            "kci_stage": kci_stage,
            "_source": "KCI",
            "_key": norm_name(site),
        })
    return rows


def load_aemo_map_names() -> set[str]:
    """Return normalized site-name keys that appear on any AEMO map PDF."""
    p = INTERMEDIATE / "aemo_pdf_matches.json"
    if not p.exists():
        return set()
    data = json.loads(p.read_text(encoding="utf-8"))
    keys: set[str] = set()
    for _, info in data.items():
        for m in info.get("matches", []):
            keys.add(norm_name(m["matched_site"]))
    return keys


def merge(nem: list[dict], kci: list[dict]) -> list[dict]:
    """Merge by normalized site name. Output one row per project with combined data."""
    nem_by_key: dict[str, list[dict]] = {}
    for r in nem:
        nem_by_key.setdefault(r["_key"], []).append(r)

    kci_by_key: dict[str, list[dict]] = {}
    for r in kci:
        kci_by_key.setdefault(r["_key"], []).append(r)

    aemo_keys = load_aemo_map_names()
    merged: list[dict] = []
    seen_keys: set[str] = set()

    # iterate NEM first (authoritative for capacity / DUID)
    for key, group in nem_by_key.items():
        if not key:
            continue
        seen_keys.add(key)
        primary = max(group, key=lambda r: r.get("capacity_mw") or 0)
        total_cap = sum((r.get("capacity_mw") or 0) for r in group)
        kci_match = kci_by_key.get(key, [])
        loc = next((k["location_desc"] for k in kci_match if k.get("location_desc")), "")
        on_aemo = key in aemo_keys

        stage = classify_stage(primary, in_kci=bool(kci_match), on_aemo_map=on_aemo)
        merged.append({
            "site_name": primary["site_name"],
            "region": primary["region"],
            "state": primary["state"],
            "owner": primary["owner"],
            "technology": primary["technology"],
            "fuel": primary["fuel"],
            "capacity_mw": round(total_cap, 2) if total_cap else primary["capacity_mw"],
            "storage_mwh": primary.get("storage_mwh"),
            "unit_status": primary["unit_status"],
            "asset_type": primary["asset_type"],
            "location_desc": loc,
            "stage": stage,
            "source": "NEM+KCI" if kci_match else "NEM",
            "on_aemo_map": on_aemo,
            "duid": primary.get("duid", ""),
        })

    # KCI-only rows -> Enquiry by default (Application if confirmed on AEMO map)
    for key, group in kci_by_key.items():
        if not key or key in seen_keys:
            continue
        primary = max(group, key=lambda r: r.get("capacity_mw") or 0)
        on_aemo = key in aemo_keys
        # KCI-only & on AEMO map => brief says this means "Application" stage was named on the map
        stage = "Application" if on_aemo else "Enquiry"
        merged.append({
            "site_name": primary["site_name"],
            "region": primary["region"],
            "state": primary["state"],
            "owner": primary["owner"],
            "technology": primary["technology"],
            "fuel": "",
            "capacity_mw": primary["capacity_mw"],
            "storage_mwh": None,
            "unit_status": "",
            "asset_type": "Project",
            "location_desc": primary.get("location_desc", ""),
            "stage": stage,
            "source": "KCI",
            "on_aemo_map": on_aemo,
            "duid": "",
        })

    return merged


def classify_stage(nem_row: dict, in_kci: bool, on_aemo_map: bool = False) -> str:
    """Map NEM Gen Info Unit Status (+ AEMO map presence) to a display stage.

    Rules:
      In Service / Announced Withdrawal -> Existing
      Committed / In Commissioning      -> Committed
      Anticipated                       -> Anticipated (upgrade to Committed if on AEMO map)
      Withdrawn - Permanent             -> Withdrawn
      Publicly Announced + on AEMO map  -> Committed   (named on Regional Boundaries layer)
      Publicly Announced (otherwise)    -> Application
    """
    us = (nem_row.get("unit_status") or "").lower()
    if us.startswith("in service"):
        return "Existing"
    if us.startswith("committed") or "commissioning" in us:
        return "Committed"
    if "anticipated" in us:
        return "Committed" if on_aemo_map else "Anticipated"
    if "withdrawn" in us:
        return "Withdrawn"
    if "publicly announced" in us or us == "proposed":
        return "Committed" if on_aemo_map else "Application"
    return "Unknown"


def assign_coords(rows: list[dict], seed: int = 42) -> None:
    """Region-centroid + deterministic jitter. Placeholder for real geocoding.

    Future hook: replace this with a geocoder that uses `location_desc`.
    """
    rng = random.Random(seed)
    for r in rows:
        region = r.get("region")
        if region not in REGION_CENTROID:
            r["lat"], r["lon"] = None, None
            continue
        lat0, lon0 = REGION_CENTROID[region]
        dlat, dlon = REGION_JITTER[region]
        # seed jitter on site name so reruns are stable per project
        seed_n = abs(hash(r["site_name"])) % (10**8)
        rr = random.Random(seed_n)
        r["lat"] = lat0 + (rr.random() - 0.5) * 2 * dlat
        r["lon"] = lon0 + (rr.random() - 0.5) * 2 * dlon


def build_map(rows: list[dict], out_html: Path) -> None:
    import folium
    from folium.plugins import MarkerCluster

    m = folium.Map(location=[-30.0, 144.0], zoom_start=5, tiles="cartodbpositron",
                   control_scale=True)

    # one FeatureGroup per stage so the user can toggle/filter
    groups: dict[str, folium.FeatureGroup] = {}
    for stage in STAGE_ORDER:
        fg = folium.FeatureGroup(name=f"{stage}", show=(stage != "Existing"))
        groups[stage] = fg

    for r in rows:
        if r["lat"] is None or r["lon"] is None:
            continue
        stage = r.get("stage", "Unknown")
        colour = STAGE_COLOUR.get(stage, STAGE_COLOUR["Unknown"])
        cap = r.get("capacity_mw") or 0
        # radius: log-scale, clamped
        radius = max(3, min(18, 3 + (math.log10(cap + 1) * 4))) if cap else 3
        popup_html = f"""
        <div style='font-family:system-ui;font-size:12px;min-width:220px'>
          <div style='font-weight:600;font-size:13px;margin-bottom:4px'>{r['site_name']}</div>
          <div><b>Stage:</b> <span style='color:{colour}'>{stage}</span></div>
          <div><b>Capacity:</b> {cap:.1f} MW{(' / ' + str(r['storage_mwh']) + ' MWh') if r.get('storage_mwh') else ''}</div>
          <div><b>Region:</b> {r['region']} ({r['state']})</div>
          <div><b>Technology:</b> {r.get('technology','')}</div>
          <div><b>Owner:</b> {r.get('owner','')}</div>
          <div><b>Location:</b> {r.get('location_desc','') or '(approx region centroid)'}</div>
          <div style='color:#888;margin-top:4px'>Source: {r.get('source','')}{' • on AEMO map' if r.get('on_aemo_map') else ''}</div>
        </div>
        """
        folium.CircleMarker(
            location=[r["lat"], r["lon"]],
            radius=radius,
            color=colour,
            weight=1,
            fill=True,
            fill_color=colour,
            fill_opacity=0.75,
            tooltip=f"{r['site_name']} — {cap:.0f} MW — {stage}",
            popup=folium.Popup(popup_html, max_width=300),
        ).add_to(groups[stage])

    for stage in STAGE_ORDER:
        groups[stage].add_to(m)

    folium.LayerControl(collapsed=False).add_to(m)

    # Legend
    legend = """
    <div style='position:fixed;bottom:24px;left:24px;z-index:9999;background:white;
                padding:10px 14px;border:1px solid #ccc;border-radius:6px;
                font-family:system-ui;font-size:12px;box-shadow:0 1px 3px rgba(0,0,0,.15)'>
      <div style='font-weight:600;margin-bottom:6px'>Project stage</div>
    """
    for stage in STAGE_ORDER:
        legend += (
            f"<div style='display:flex;align-items:center;margin:2px 0'>"
            f"<span style='display:inline-block;width:12px;height:12px;border-radius:50%;"
            f"background:{STAGE_COLOUR[stage]};margin-right:6px'></span>{stage}</div>"
        )
    legend += "<div style='margin-top:6px;color:#888;font-size:10px'>Marker size ∝ log(capacity)</div></div>"
    m.get_root().html.add_child(folium.Element(legend))

    out_html.write_text(m.get_root().render(), encoding="utf-8")


def main() -> None:
    nem = read_nem_gen_info()
    kci = read_kci()
    merged = merge(nem, kci)
    assign_coords(merged)

    # Stats
    from collections import Counter
    stage_ct = Counter(r["stage"] for r in merged)
    source_ct = Counter(r["source"] for r in merged)
    region_ct = Counter(r["region"] for r in merged)
    print(f"Total projects: {len(merged)}")
    print(f"By stage: {dict(stage_ct)}")
    print(f"By source: {dict(source_ct)}")
    print(f"By region: {dict(region_ct)}")

    (INTERMEDIATE / "projects.json").write_text(
        json.dumps(merged, indent=2, default=str), encoding="utf-8"
    )
    print(f"Wrote {INTERMEDIATE/'projects.json'}")

    # Folium output retained as fallback; canonical viewer is outputs/nem_map.html
    build_map(merged, OUTPUTS / "nem_generation_map_folium.html")
    print(f"Wrote {OUTPUTS/'nem_generation_map_folium.html'}")


if __name__ == "__main__":
    main()
